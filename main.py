import tkinter as tk
from tkinter import messagebox
from datetime import datetime
import threading
from openpyxl.styles import PatternFill

def generate_excel(user_id, password, search_date, progress_callback):
    import requests
    from bs4 import BeautifulSoup
    import openpyxl
    import re
    import os

    with open("발전소이름.txt", encoding="utf-8") as f:
        site_names = [line.strip() for line in f if line.strip()]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "발전량 통계"
    for idx, name in enumerate(site_names, start=2):
        ws.cell(row=idx, column=1, value=name)
    site_name_row = {name: idx for idx, name in enumerate(site_names, start=2)}
    next_row = len(site_names) + 2

    session = requests.Session()
    login_url = "http://iplug.dasstech.com/loginRequest"
    login_data = {
        "id": user_id,
        "pass": password
    }
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36",
        "Referer": "http://iplug.dasstech.com/login",
        "Origin": "http://iplug.dasstech.com"
    }
    session.post(login_url, data=login_data, headers=headers)

    def clean_name(name):
        name = re.sub(r'(태양광발전소|태양광 발전소|태양광)$', '', name).strip()
        return name.replace(" ", "")

    site_info = []
    total_pages = 14
    for page in range(1, total_pages + 1):
        url = f"http://iplug.dasstech.com/monitoring/getProjectList.json?currentPage={page}"
        res = session.get(url)
        data = res.json()
        for proj in data.get("list", []):
            raw_name = proj["SITE_NAME"]
            cleaned = clean_name(raw_name)
            site_info.append({
                "SITE_CODE": proj["SITE_CODE"],
                "SITE_NAME": cleaned
            })
        # 진행률 업데이트 (페이지 기준)
        percent = int(page / total_pages * 100)
        progress_callback(f"프로젝트 목록 수집: {percent}%")

    search_unit = "month"
    date_set = set()
    result = {}

    for idx, site in enumerate(site_info):
        site_code = site["SITE_CODE"]
        site_name = site["SITE_NAME"]
        url = f"http://iplug.dasstech.com/monitoring/dataSearch/netgenerationstats?SITE_CODE={site_code}&SEARCH_DATE={search_date}&CALC_PRICE=&SEARCH_UNIT={search_unit}"
        resp = session.get(url)
        soup = BeautifulSoup(resp.text, "html.parser")
        rows = soup.select("table.tstyle1 tbody tr")
        if site_name not in result:
            result[site_name] = {}
        for row in rows:
            cols = row.find_all("td")
            if len(cols) >= 2:
                date = cols[0].get_text(strip=True)
                gen = cols[1].get_text(strip=True)
                result[site_name][date] = gen
                date_set.add(date)
        # 진행률 업데이트 (사이트별)
        percent = int((idx + 1) / len(site_info) * 100)
        progress_callback(f"발전량 수집: {percent}%")

    date_list = sorted(date_set)
    for idx, date in enumerate(date_list, start=2):
        date_formatted = date.replace('.', '-')
        ws.cell(row=1, column=idx, value=date_formatted)

    green_fill = PatternFill(start_color="CCFF99", end_color="CCFF99", fill_type="solid")

    added_blank = False
    for site_name, data in result.items():
        row_idx = site_name_row.get(site_name)
        if not row_idx:
            if not added_blank:
                # 기존 그룹과 새 그룹 사이에 한 줄 띄우기
                next_row += 1
                added_blank = True
            row_idx = next_row
            ws.cell(row=row_idx, column=1, value=site_name)
            ws.cell(row=row_idx, column=1).fill = green_fill  # 연두색 칠하기
            site_name_row[site_name] = row_idx
            next_row += 1
        for col_idx, date in enumerate(date_list, start=2):
            ws.cell(row=row_idx, column=col_idx, value=data.get(date, ""))

    base = "generation_stats"
    ext = ".xlsx"
    i = 1
    filename = f"{base}{ext}"
    while os.path.exists(filename):
        filename = f"{base}{i}{ext}"
        i += 1

    wb.save(filename)
    return filename

def run_gui():
    def on_generate():
        user_id = entry_id.get()
        password = entry_pw.get()
        search_date = entry_date.get()
        btn["state"] = "disabled"
        progress_label["text"] = "진행중..."
        def task():
            try:
                filename = generate_excel(
                    user_id, password, search_date,
                    lambda msg: progress_label.after(0, progress_label.config, {"text": msg})
                )
                progress_label.after(0, progress_label.config, {"text": "완료!"})
                messagebox.showinfo("완료", f"엑셀 저장 완료: {filename}")
            except Exception as e:
                progress_label.after(0, progress_label.config, {"text": "오류 발생"})
                messagebox.showerror("오류", str(e))
            finally:
                btn["state"] = "normal"
        threading.Thread(target=task).start()

    root = tk.Tk()
    root.title("발전량 엑셀 생성기")
    root.geometry("350x260")

    tk.Label(root, text="아이디:").grid(row=0, column=0, padx=10, pady=10, sticky="e")
    entry_id = tk.Entry(root)
    entry_id.grid(row=0, column=1, padx=10, pady=10)

    tk.Label(root, text="비밀번호:").grid(row=1, column=0, padx=10, pady=10, sticky="e")
    entry_pw = tk.Entry(root, show="*")
    entry_pw.grid(row=1, column=1, padx=10, pady=10)

    tk.Label(root, text="날짜(YYYY.MM.DD):").grid(row=2, column=0, padx=10, pady=10, sticky="e")
    entry_date = tk.Entry(root)
    today_str = datetime.now().strftime("%Y.%m.%d")
    entry_date.insert(0, today_str)
    entry_date.grid(row=2, column=1, padx=10, pady=10)

    btn = tk.Button(root, text="엑셀 생성", command=on_generate)
    btn.grid(row=3, column=0, columnspan=2, pady=10)

    progress_label = tk.Label(root, text="")
    progress_label.grid(row=4, column=0, columnspan=2, pady=5)

    root.mainloop()

if __name__ == "__main__":
    run_gui()