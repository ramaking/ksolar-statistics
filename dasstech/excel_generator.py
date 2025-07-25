import os
import re
from datetime import datetime
from openpyxl.styles import PatternFill
from web_crawler import login, get_project_list, get_generation_stats

def clean_name(name):
    name = re.sub(r'(태양광발전소|태양광 발전소|태양광)$', '', name).strip()
    return name.replace(" ", "")

def read_exclude_names(filename="제외목록.txt"):
    if not os.path.exists(filename):
        return set()
    with open(filename, encoding="utf-8") as f:
        return set(line.strip() for line in f if line.strip())

def read_site_names(exclude_names, filename="발전소이름.txt"):
    if not os.path.exists(filename):
        return []
    with open(filename, encoding="utf-8") as f:
        return [line.strip() for line in f if line.strip() and line.strip() not in exclude_names]

def generate_excel(user_id, password, search_date, progress_callback):
    import openpyxl

    exclude_names = read_exclude_names()
    site_names = read_site_names(exclude_names)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "발전량 통계"
    for idx, name in enumerate(site_names, start=2):
        ws.cell(row=idx, column=2, value=name)
    site_name_row = {name: idx for idx, name in enumerate(site_names, start=2)}
    next_row = len(site_names) + 2

    # 로그인 및 세션 획득
    session = login(user_id, password)

    # 프로젝트 리스트 수집
    site_info = []
    total_pages = 14
    for page in range(1, total_pages + 1):
        projects = get_project_list(session, page)
        for proj in projects:
            raw_name = proj["SITE_NAME"]
            cleaned = clean_name(raw_name)
            if cleaned in exclude_names:
                continue
            site_info.append({
                "SITE_CODE": proj["SITE_CODE"],
                "SITE_NAME": cleaned
            })
        percent = int(page / total_pages * 100)
        progress_callback(f"프로젝트 목록 수집: {percent}%")

    # 발전량 데이터 수집
    search_unit = "month"
    date_set = set()
    result = {}
    for idx, site in enumerate(site_info):
        site_code = site["SITE_CODE"]
        site_name = site["SITE_NAME"]
        stats = get_generation_stats(session, site_code, search_date, search_unit)
        if site_name not in result:
            result[site_name] = {}
        for date, gen in stats.items():
            result[site_name][date] = gen
            date_set.add(date)
        percent = int((idx + 1) / len(site_info) * 100)
        progress_callback(f"발전량 수집: {percent}%")

    date_list = sorted(date_set)
    today_str = datetime.now().strftime("%Y-%m-%d").replace("-", ".")

    ws.cell(row=1, column=1, value="링크")
    ws.cell(row=1, column=2, value="SITE_NAME")
    for idx, date in enumerate(date_list, start=3):
        date_formatted = date.replace('.', '-')
        ws.cell(row=1, column=idx, value=date_formatted)

    green_fill = PatternFill(start_color="CCFF99", end_color="CCFF99", fill_type="solid")
    added_blank = False

    for site_name, data in result.items():
        row_idx = site_name_row.get(site_name)
        site_code = next((s["SITE_CODE"] for s in site_info if s["SITE_NAME"] == site_name), None)
        if not row_idx:
            if not added_blank:
                next_row += 1
                added_blank = True
            row_idx = next_row
            ws.cell(row=row_idx, column=2, value=site_name)
            ws.cell(row=row_idx, column=2).fill = green_fill
            site_name_row[site_name] = row_idx
            next_row += 1
        else:
            ws.cell(row=row_idx, column=2, value=site_name)
        if site_code:
            url = f"http://iplug.dasstech.com/monitoring/dataSearch/netgenerationstats?SITE_CODE={site_code}&SEARCH_DATE={today_str}&CALC_PRICE=&SEARCH_UNIT=month"
            ws.cell(row=row_idx, column=1).hyperlink = url
            ws.cell(row=row_idx, column=1).value = "링크"
            ws.cell(row=row_idx, column=1).style = "Hyperlink"
        for col_idx, date in enumerate(date_list, start=3):
            ws.cell(row=row_idx, column=col_idx, value=data.get(date, ""))

    base = "generation_stats"
    ext = ".xlsx"
    i = 1
    filename = f"{base}{ext}"
    while os.path.exists(filename):
        filename = f"{base}{i}{ext}"
        i += 1

    wb.save(filename)
    return filename